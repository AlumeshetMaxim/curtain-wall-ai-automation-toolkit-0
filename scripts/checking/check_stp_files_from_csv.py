"""Check which items from a CSV exist as STP/STEP files in a folder.

Creates an Excel report and highlights missing items.
"""

from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

STP_FOLDER = Path("./examples/stp")
CSV_FILE = Path("./examples/items.csv")
OUTPUT_XLSX = Path("./reports/stp_check.xlsx")
FILENAME_COLUMN = None
RECURSIVE = True


def normalize_name(value: str) -> str:
    value = (value or "").strip().lower().replace("\\", "/").split("/")[-1]
    value = re.sub(r"\s+", " ", value)
    for suffix in [".stp", ".step"]:
        if value.endswith(suffix):
            value = value[: -len(suffix)]
    return value


def build_stp_index(folder: Path, recursive: bool) -> set[str]:
    pattern = "**/*" if recursive else "*"
    return {normalize_name(path.stem) for path in folder.glob(pattern) if path.suffix.lower() in {".stp", ".step"}}


def auto_detect_filename_column(df: pd.DataFrame) -> str:
    preferred = ["filename", "file", "stp", "step", "model", "name", "item", "part"]
    for key in preferred:
        for column in df.columns:
            if key in str(column).lower():
                return column
    return df.columns[0]


def main() -> None:
    df = pd.read_csv(CSV_FILE)
    column = FILENAME_COLUMN or auto_detect_filename_column(df)
    existing = build_stp_index(STP_FOLDER, RECURSIVE)

    df["NORMALIZED_NAME"] = [normalize_name(str(value)) for value in df[column]]
    df["EXISTS_IN_STP_FOLDER"] = df["NORMALIZED_NAME"].apply(lambda value: value in existing)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_XLSX, index=False)

    workbook = load_workbook(OUTPUT_XLSX)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    exists_col = header.index("EXISTS_IN_STP_FOLDER") + 1
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    red_font = Font(color="FF9C0006")

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=exists_col).value is False:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = red_fill
                sheet.cell(row=row, column=col).font = red_font

    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    workbook.save(OUTPUT_XLSX)
    print(f"Created: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
